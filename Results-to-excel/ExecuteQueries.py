import pandas as pd
# import numpy as np
from FormatQueries import *
from sqlalchemy import create_engine
from openpyxl import Workbook, load_workbook
# from openpyxl.styles import Border, Alignment, PatternFill

def get_database():
    print("\nAvailable databases: ")
    print("---------------")
    print("module2") #mod2, mod12
    print("classicmodels") #module6, mod10, mod11
    print("employees") #project1
    print("sakila") #midterm proj
    print("project2\n") #proj2, final proj
    database = input("What database are you using? ")

    if database != 'module2' and database != 'classicmodels' and database != 'employees' and database != 'sakila' and database != 'project2':
        print(f"{database} does not exist in the avaliable databases")
        return get_database()
    return database

def create_connector_engine(database_name):
    # mydb = mysql.connector.connect(
    #     host="mysql-container",
    #     user="root",
    #     passwd="root",
    #     database=f"{database_name}"
    # )
    engine_uri = f"mysql+mysqlconnector://root:root@mysql-container:3306/{database_name}"
    engine = create_engine(engine_uri)

    return engine

def create_workbook(path):
   workbook = Workbook()
   workbook.save(path) 

def execute_formatted_file(database, path):
    mydb = create_connector_engine(database)
    with open ('FormattedQueries.sql') as file:
        queries_file = file.read()
        queries = queries_file.splitlines()
        page = 1
        for q in queries:
            # if line in file is empty line, then skip
            if (len(q.strip()) == 0):
                continue
            pd_query = pd.read_sql(q, mydb)
            try:
                df = pd.DataFrame(pd_query)
                with pd.ExcelWriter(path,mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer:
                    # df.to_excel(writer,sheet_name="Sheet",startrow=writer.sheets["Sheet"].max_row,index=False)
                    df.to_excel(writer,sheet_name=f"Sheet{page}", index=False)
                    print(f'Query {page} write successfully')
                page += 1
            except Exception as err:
                print(f"Error Occured: {err}")

def main():
    path = 'Result_Data.xlsx'
    writeFile()
    create_workbook(path) 
    database = get_database()
    execute_formatted_file(database, path)
    try:
        create_connector_engine(database)
        print("Successfully connected to the database!")
    except Exception as err:
        print(f"Error Occured: {err}")

main()