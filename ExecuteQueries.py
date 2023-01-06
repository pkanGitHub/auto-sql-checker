import pandas as pd
# import numpy as np
from FormatQueries import *
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.styles import Border, Alignment, PatternFill

def get_database():
    print("\nAvailable databases: ")
    print("---------------")
    print("module2") #mod2, mod12
    print("classicmodels") #module6, mod10, mod11
    print("employees") #project1
    print("sakila") #midterm proj
    print("project2\n") #proj2, final proj
    database = input("What database are you using? ")

    if database != 'module2' and database != 'classicmodels' and database != 'employees' and database != 'sakila' and database != 'project2':
        print(f"{database} does not exist in the avaliable databases")
        return get_database()
    return database

def create_connector_object(database_name):
    # mydb = mysql.connector.connect(
    #     host="mysql-container",
    #     user="root",
    #     passwd="root",
    #     database=f"{database_name}"
    # )
    engine_uri = f"mysql+mysqlconnector://root:root@mysql-container:3306/{database_name}"
    connection = create_engine(engine_uri)

    return connection

def create_workbook(path):
   workbook = Workbook()
   workbook.save(path) 

def cell_style(path):
    workbook = Workbook()
    sheet = workbook.active
    # for wrapRows in sheet.iter_rows(min_row=1, min_col=1):
    #     for cell in wrapRows:
    #         cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    sheet.column_dimensions

    # green = "96C8A2"
    # for colorRow in sheet.head(1):
    #     for colorCell in colorRow:
    #         colorCell.fill = PatternFill(start_color=green, end_color=green,fill_type = "solid")
    workbook.save(path)


def read_formatted_file(database, path):
    mydb = create_connector_object(database)
    with open ('FormattedQueries.sql') as file:
        queries_file = file.read()
        queries = queries_file.splitlines()
        page = 1
        for q in queries:
            pd_query = pd.read_sql(q, mydb)
            try:
                df = pd.DataFrame(pd_query)
                with pd.ExcelWriter(path,mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer:
                    # df.to_excel(writer,sheet_name="Sheet",startrow=writer.sheets["Sheet"].max_row,index=False)
                    df.to_excel(writer,sheet_name=f"Sheet{page}", index=False)
                    print(f'Query {page} write successfully')
                    page += 1
            except Exception as err:
                print(f"Error Occured: {err}")

def main():
    path = 'Result_Data.xlsx'
    writeFile()
    create_workbook(path) 
    database = get_database()
    read_formatted_file(database, path)
    try:
        create_connector_object(database)
        print("Successfully connected to the database!")
    except Exception as err:
        print(f"Error Occured: {err}")

main()