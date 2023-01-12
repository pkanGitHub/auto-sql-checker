import pandas as pd
from FormatQueries import *
from FormatSolutionQueries import *
from sqlalchemy import create_engine
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter

# from openpyxl.styles import Border, Alignment, PatternFill

def get_database():
    print("\nAvailable databases: ")
    print("---------------")
    print("module2") #mod2, mod12
    print("classicmodels") #module6, mod10, mod11
    print("employees") #project1
    print("sakila") #midterm proj
    print("project2\n") #proj2, final proj
    database = input("What database are you using? ")

    if database != 'module2' and database != 'classicmodels' and database != 'employees' and database != 'sakila' and database != 'project2':
        print(f"{database} does not exist in the avaliable databases")
        return get_database()
    return database

def create_connector_engine(database_name):
    engine_uri = f"mysql+mysqlconnector://root:root@mysql-container:3306/{database_name}"
    engine = create_engine(engine_uri)
    return engine

#create empty excel file
def create_workbook(path):
   workbook = Workbook()
   workbook.save(path) 

def execute_formatted_file(database, path, formatted_file):
    mydb = create_connector_engine(database)
    with open (formatted_file) as file:
        queries_file = file.read()
        queries = queries_file.splitlines()
        page = 1
        for q in queries:
            # if line in file is empty, then skip
            if (len(q.strip()) == 0):
                continue
            pd_query = pd.read_sql(q, mydb)
            try:
                df = pd.DataFrame(pd_query)
                with pd.ExcelWriter(path,mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer:
                    # df.to_excel(writer,sheet_name="Sheet",startrow=writer.sheets["Sheet"].max_row,index=False)
                    df.to_excel(writer,sheet_name=f"Sheet{page}", index=False)
                    if formatted_file == 'FormattedQueries.sql':
                        print(f'Query {page} write successfully')
                page += 1
            except Exception as err:
                print(f"Error Occured: {err}")

def remove_default_sheet(path):
    wb = load_workbook(path)
    del wb['Sheet']
    wb.save(path)

def auto_adjust_column_width(path):
    wb = load_workbook(path)
    sheets_dict = pd.read_excel(path, sheet_name=None)
    all_sheets = []
    for name, sheet in sheets_dict.items():
        sheet['sheet'] = name
        all_sheets.append(name)
    # iterate thru every sheets
    for sh in all_sheets:
        ws = wb[sh]
        # get into each col in the sheet
        for col in ws.columns:
            max_length = 0
            column = col[0].column #this will get the column number
            letter = get_column_letter(column) # convert column number into column letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                    adjusted_width = max_length * 1.6
                except:
                    pass
                ws.column_dimensions[letter].width = adjusted_width
    wb.save(path)

def main():
    database = get_database()
    files = {'FormattedQueries.sql': 'Result_Data.xlsx', 'FormattedSolution.sql': 'Solution_Result.xlsx'}
    try:
        create_connector_engine(database)
        print(f"{database} connected, now processing...\n")
        writeFile()
        writeSolutionFile()
        for file, path in files.items():
            create_workbook(path) 
            execute_formatted_file(database, path, file)
            remove_default_sheet(path)
            auto_adjust_column_width(path)
    except Exception as err:
        print(f"Error Occured: {err}")

main()